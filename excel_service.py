import os
import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.worksheet.worksheet import Worksheet

from numpy import ndarray, bool_


def _create_file_path() -> str:
    """
    This function creates a free file path with 'results_xslx_<no.>' format.
    """
    results_directory = "results"
    if not os.path.exists(results_directory):
        os.makedirs(results_directory)

    i = 1
    file_path = f"results/results_xslx_{i}.xlsx"
    while os.path.exists(file_path):
        i += 1
        file_path = f"results/results_xslx_{i}.xlsx"

    return file_path


def _create_freq_vs_fftmagnitude_plot(ws: Worksheet, frequencies: ndarray[float], mask: ndarray[bool_]) -> None:
    """
    This function creates a Frequency vs FFT Magnitude plot.
    """
    chart = ScatterChart()
    x_values = Reference(ws, min_col=1, min_row=2, max_row=len(frequencies[mask]) + 1)
    y_values = Reference(ws, min_col=2, min_row=1, max_row=len(frequencies[mask]) + 1)
    series = Series(y_values, x_values, title_from_data=True)
    chart.series.append(series)
    chart.x_axis.title = "Frequency [Hz]"
    chart.y_axis.title = "Amplitude [db]"
    ws.add_chart(chart, "E4")


def create_excel_file(frequencies: ndarray[float], fft_abs: ndarray[float], mask: ndarray[bool_]) -> None:
    """
    This function save Frequency, FFT Magnitude results and Frequency vs FFT Magnitude plot into .xsls file.
    """

    file_path = _create_file_path()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Frequency (Hz)"
    ws["B1"] = "FFT Magnitude"
    for i, frequency in enumerate(frequencies[mask]):
        ws.cell(row=i + 2, column=1, value=frequency)
        ws.cell(row=i + 2, column=2, value=fft_abs[mask][i])

    _create_freq_vs_fftmagnitude_plot(ws, frequencies, mask)

    try:
        wb.save(file_path)
    except PermissionError:
        print("You have to close the file before saving")
